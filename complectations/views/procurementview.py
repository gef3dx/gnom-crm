import os
import uuid
from decimal import Decimal
from typing import List
from PIL import Image
from django.conf import settings
from django.contrib import messages
from django.db.models import Sum
import datetime
from django.http import HttpResponse, Http404
from openpyxl.styles import Alignment, Font
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.platypus import Paragraph, Spacer
from reportlab.platypus import Image as Img
from django.views import View
from complectations.utils import (
    setup_pdf,
    setup_styles,
    create_table,
    get_filename, compress_image,
)
import io
from django.urls import reverse
from django.shortcuts import get_object_or_404, redirect
from django.views.generic import (
    ListView,
    UpdateView,
    CreateView,
    DeleteView
)
from complectations.models import (
    Complectation,
    Procurement,
)
from complectations.forms import (
    ProcurementForm
)
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from urllib.parse import quote
from django.core.files.uploadedfile import InMemoryUploadedFile


class ProcurementViewList(ListView):
    """View вывода закупок"""
    template_name = 'complectations/procurement/procurements.html'
    model = Procurement
    context_object_name = 'procurements'

    def get_queryset(self):
        slug = self.kwargs['slug']
        # Предзагружаем complectation, чтобы избежать лишних запросов при доступе
        return Procurement.objects.filter(complectation__slug=slug).select_related('complectation')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        complectation = get_object_or_404(Complectation, slug=self.kwargs['slug'])

        result = self.object_list.aggregate(
            total_price=Sum('price'),
            total_price_procent=Sum('price_procent')
        )
        total_sum = (result['total_price'] or 0) + (result['total_price_procent'] or 0)

        context.update({
            'comp_id': complectation.id,
            'comp_name': complectation.name,
            'comp_slug': complectation.slug,
            # Вычисляем сумму и процент и добавляем их в контекст
            'sum': Decimal(str(result['total_price'])).quantize(Decimal("1.00")) if result.get('total_price') else Decimal("0.00"),
            'sum_procent': Decimal(str(result['total_price_procent'])).quantize(Decimal("1.00")) if result.get('total_price_procent') else Decimal("0.00"),
            'sum_and_procent': Decimal(str(total_sum)).quantize(Decimal("1.00")) if total_sum else Decimal("0.00")
        })

        return context

    def post(self, request, *args, **kwargs):
        # Получаем список ID из POST-запроса
        procurement_ids = request.POST.getlist('procurement_ids')

        if procurement_ids:
            # Обновляем записи с указанными ID
            updated_count = Procurement.objects.filter(id__in=procurement_ids).update(payment_status="Оплатил")

            # Сообщение об успешном обновлении
            messages.success(request, f"Статус {updated_count} закупок изменен.")
        else:
            # Сообщение, если ничего не выбрано
            messages.error(request, "Выберите хотя бы одну запись для обновления.")

        # Переадресация обратно на страницу закупок
        return redirect(reverse('procurementslistpage', kwargs={'slug': self.kwargs['slug']}))


class ProcurementViewPaidList(ListView):
    """View вывода оплаченных закупок"""
    template_name = 'complectations/procurement/procurements.html'
    model = Procurement
    context_object_name = 'procurements'

    def get_queryset(self):
        slug = self.kwargs['slug']
        queryset = Procurement.objects.filter(
            complectation__slug=slug,
            payment_status="Оплатил"
        ).select_related('complectation')

        # Если записей нет, выполняем переадресацию
        if not queryset.exists():
            return redirect(reverse('procurementslistpage', kwargs={'slug': self.kwargs['slug']}))

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        complectation = get_object_or_404(Complectation, slug=self.kwargs['slug'])

        result = self.object_list.aggregate(
            total_price=Sum('price'),
            total_price_procent=Sum('price_procent')
        )
        total_sum = (result['total_price'] or 0) + (result['total_price_procent'] or 0)

        context.update({
            'comp_id': complectation.id,
            'comp_name': complectation.name,
            'comp_slug': complectation.slug,
            # Вычисляем сумму и процент и добавляем их в контекст
            'sum': Decimal(str(result['total_price'])).quantize(Decimal("1.00")) if result.get('total_price') else Decimal("0.00"),
            'sum_procent': Decimal(str(result['total_price_procent'])).quantize(Decimal("1.00")) if result.get('total_price_procent') else Decimal("0.00"),
            'sum_and_procent': Decimal(str(total_sum)).quantize(Decimal("1.00")) if total_sum else Decimal("0.00")
        })

        return context

    def post(self, request, *args, **kwargs):
        # Получаем список ID из POST-запроса
        procurement_ids = request.POST.getlist('procurement_ids')

        if procurement_ids:
            # Обновляем записи с указанными ID
            updated_count = Procurement.objects.filter(id__in=procurement_ids).update(payment_status="Оплатил")

            # Сообщение об успешном обновлении
            messages.success(request, f"Статус {updated_count} закупок изменен.")
        else:
            # Сообщение, если ничего не выбрано
            messages.warning(request, "Выберите хотя бы одну запись для обновления.")

        # Переадресация обратно на страницу закупок
        return redirect(reverse('procurementslistpage', kwargs={'slug': self.kwargs['slug']}))


class ProcurementViewUnpaidList(ListView):
    """View вывода не оплаченных закупок"""
    template_name = 'complectations/procurement/procurements.html'
    model = Procurement
    context_object_name = 'procurements'

    def get_queryset(self):
        slug = self.kwargs['slug']
        # Фильтруем по `complectation__slug` и `payment_status`
        return Procurement.objects.filter(
            complectation__slug=slug,
            payment_status="Не оплатил"
        ).select_related('complectation')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        complectation = get_object_or_404(Complectation, slug=self.kwargs['slug'])

        result = self.object_list.aggregate(
            total_price=Sum('price'),
            total_price_procent=Sum('price_procent')
        )
        total_sum = (result['total_price'] or 0) + (result['total_price_procent'] or 0)

        context.update({
            'comp_id': complectation.id,
            'comp_name': complectation.name,
            'comp_slug': complectation.slug,
            # Вычисляем сумму и процент и добавляем их в контекст
            'sum': Decimal(str(result['total_price'])).quantize(Decimal("1.00")) if result.get('total_price') else Decimal("0.00"),
            'sum_procent': Decimal(str(result['total_price_procent'])).quantize(Decimal("1.00")) if result.get('total_price_procent') else Decimal("0.00"),
            'sum_and_procent': Decimal(str(total_sum)).quantize(Decimal("1.00")) if total_sum else Decimal("0.00")
        })

        return context

    def post(self, request, *args, **kwargs):
        # Получаем список ID из POST-запроса
        procurement_ids = request.POST.getlist('procurement_ids')

        if procurement_ids:
            # Обновляем записи с указанными ID
            updated_count = Procurement.objects.filter(id__in=procurement_ids).update(payment_status="Оплатил")

            # Сообщение об успешном обновлении
            messages.success(request, f"Статус {updated_count} закупок изменен.")
        else:
            # Сообщение, если ничего не выбрано
            messages.warning(request, "Выберите хотя бы одну запись для обновления.")

        # Переадресация обратно на страницу закупок
        return redirect(reverse('procurementslistpage', kwargs={'slug': self.kwargs['slug']}))


class ProcurementViewAdd(CreateView):
    """View добавления закупок"""
    template_name = "complectations/procurement/procurementsadd.html"
    model = Procurement
    form_class = ProcurementForm

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.author = self.request.user
        obj.price_procent = obj.price * (Decimal(obj.procent / 100))
        # Получаем объект Complectation и сохраняем его как атрибут класса
        self.complectation = get_object_or_404(Complectation, id=self.kwargs['product_id'])
        obj.complectation = self.complectation

        xlsx_field = self.request.FILES.get('xlsx_file')

        if xlsx_field:
            allowed_extensions = ['xlsx', 'xlsm', 'xlsb', 'xltx']
            file_extension = xlsx_field.name.split('.')[-1].lower()
            if file_extension not in allowed_extensions:
                form.add_error('xlsx_file', "Только файлы с расширениями: xlsx, xlsm, xlsb, xltx допустимы.")
                return self.form_invalid(form)

        image_field = self.request.FILES.get('image')

        if image_field:
            allowed_extensions = ['jpg', 'jpeg', 'png']
            file_extension = image_field.name.split('.')[-1].lower()
            if file_extension not in allowed_extensions:
                form.add_error('image', "Только файлы с расширениями: jpg, jpeg, png допустимы.")
                return self.form_invalid(form)

            if image_field.size > 4 * 1024 * 1024:
                form.add_error('image', "Размер изображения не должен превышать 4 МБ.")
                return self.form_invalid(form)

            # Открываем и сжимаем изображение
            img = Image.open(image_field)
            compressed_image = compress_image(img)

            # Генерируем уникальное имя для файла
            unique_filename = f"img_{uuid.uuid4().hex}.jpg"

            # Определяем путь для сохранения изображения
            slug_folder = os.path.join('procurement', self.complectation.slug)  # относительный путь
            full_path = os.path.join(settings.MEDIA_ROOT, slug_folder)  # абсолютный путь
            os.makedirs(full_path, exist_ok=True)  # Создаем папку, если ее нет

            # Сохраняем изображение в нужной папке
            output = io.BytesIO()
            compressed_image.save(output, format='JPEG')
            output.seek(0)

            # Задаем относительный путь для поля image
            obj.image = InMemoryUploadedFile(
                output, 'ImageField', os.path.join(slug_folder, unique_filename),
                'image/jpeg', output.tell(), None
            )

        obj.save()  # Сохраняем объект Procurement
        messages.success(self.request, "Услуга добавлена")
        return super().form_valid(form)

    def get_success_url(self):
        # Используем сохраненный атрибут self.complectation, чтобы избежать лишнего запроса
        return reverse('procurementslistpage', kwargs={'slug': self.complectation.slug})


class ProcurementViewEdit(UpdateView):
    """View редактирования закупок"""
    template_name = "complectations/procurement/procurementsedit.html"
    model = Procurement
    form_class = ProcurementForm

    def form_valid(self, form):
        # Сохраняем объект в self.object для доступа в get_success_url
        self.object = form.save(commit=False)
        self.object.author = self.request.user
        self.object.price_procent = self.object.price * (Decimal(self.object.procent) / 100)

        # Получаем complectation из объекта
        self.complectation = self.object.complectation  # Убедитесь, что у вас есть это поле в модели

        xlsx_field = self.request.FILES.get('xlsx_file')

        if xlsx_field:
            allowed_extensions = ['xlsx', 'xlsm', 'xlsb', 'xltx']
            file_extension = xlsx_field.name.split('.')[-1].lower()
            if file_extension not in allowed_extensions:
                form.add_error('xlsx_file', "Только файлы с расширениями: xlsx, xlsm, xlsb, xltx допустимы.")
                return self.form_invalid(form)

        image_field = self.request.FILES.get('image')

        if image_field:
            allowed_extensions = ['jpg', 'jpeg', 'png']
            file_extension = image_field.name.split('.')[-1].lower()
            if file_extension not in allowed_extensions:
                form.add_error('image', "Только файлы с расширениями: jpg, jpeg, png допустимы.")
                return self.form_invalid(form)

            if image_field.size > 4 * 1024 * 1024:
                form.add_error('image', "Размер изображения не должен превышать 4 МБ.")
                return self.form_invalid(form)

            # Открываем и сжимаем изображение
            img = Image.open(image_field)
            compressed_image = compress_image(img)

            # Генерируем уникальное имя для файла
            unique_filename = f"чек_{uuid.uuid4().hex}.jpg"

            # Определяем путь для сохранения изображения
            slug_folder = os.path.join('procurement', self.complectation.slug)  # относительный путь
            full_path = os.path.join(settings.MEDIA_ROOT, slug_folder)  # абсолютный путь
            os.makedirs(full_path, exist_ok=True)  # Создаем папку, если ее нет

            # Сохраняем изображение в нужной папке
            output = io.BytesIO()
            compressed_image.save(output, format='JPEG')
            output.seek(0)

            # Задаем относительный путь для поля image
            self.object.image = InMemoryUploadedFile(
                output, 'ImageField', os.path.join(slug_folder, unique_filename),
                'image/jpeg', output.tell(), None
            )

        self.object.save()  # Сохраняем объект после обновления полей
        # Добавляем сообщение после успешного сохранения
        messages.success(self.request, "Услуга изменена")
        return super().form_valid(form)

    def get_success_url(self):
        # Используем self.object.complectation.slug для построения URL
        return reverse('procurementslistpage', kwargs={'slug': self.complectation.slug})


class ProcurementViewDelete(DeleteView):
    """View удаления закупок"""
    model = Procurement
    template_name = "complectations/procurement/procurementsdelete.html"

    def form_valid(self, form):
        messages.success(self.request, "Закупка успешно удаленна")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('procurementslistpage', kwargs={'slug': self.object.complectation.slug})


class ProcurementViewPDF(View):
    """View вывода PDF закупок"""
    def get(self, request, *args, **kwargs) -> HttpResponse:
        slug = self.kwargs['slug']
        procurements = Procurement.objects.filter(complectation__slug=slug).select_related('complectation')

        # Проверяем, что есть данные
        if not procurements.exists():
            raise Http404("Записи не найдены для данного комплектации.")

        buffer = io.BytesIO()

        # Настройка PDF и стили
        pdf = setup_pdf(buffer, size=True)
        styles = setup_styles()
        data = self.get_table_data(procurements)
        col_widths = self.calculate_column_widths(data)
        table = create_table(data, col_widths)

        # Логотип
        logo_path = os.path.join(settings.BASE_DIR, "static", "img", "gnomlogob.png")
        logo = Img(logo_path, width=2 * inch, height=1 * inch)

        # Стиль заголовка по центру
        styles.add(ParagraphStyle(name="CenteredHeading1", parent=styles["Heading1"], alignment=TA_CENTER))
        styles.add(ParagraphStyle(name="CenteredHeading2", parent=styles["Heading1"], alignment=TA_CENTER, fontSize=14))


        # Вычисление суммы и добавление в контекст
        total_sum = Decimal(procurements.aggregate(total_sum=Sum('price')).get("total_sum") or 0).quantize(Decimal("0.00"))
        total_sum_proc = Decimal(procurements.aggregate(total_sum_proc=Sum('price_procent')).get("total_sum_proc") or 0).quantize(Decimal("0.00"))

        # Элементы документа
        elements = [
            logo, Spacer(1, 0.2 * inch),
            Paragraph(f'Информация о услугах "{procurements.first().complectation}"', styles["CenteredHeading1"]),
            Paragraph(f'Сумма всех услуг: {total_sum:,.2f} р на {str(datetime.datetime.now())[:11]}'.replace(",","`").replace(".", ","), styles["CenteredHeading2"]),
            Paragraph(f'Сумма процента: {total_sum_proc:,.2f}'.replace(",", "`").replace(".", ",") + " р", styles["CenteredHeading2"]),
            Spacer(1, 0.2 * inch), table
        ]
        pdf.build(elements)

        # Возвращаем буфер к началу
        buffer.seek(0)

        # Формируем ответ
        filename = get_filename(procurements)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    def calculate_column_widths(self, data: List[List[str]]) -> List[float]:
        """Вычисляет и возвращает ширины колонок для таблицы."""
        font_name = 'FreeSans'
        font_size = 10
        number_column_width = max(stringWidth(str(row[0]), font_name, font_size) for row in data) + 10
        name_column_width = max(stringWidth(str(row[1]), font_name, font_size) for row in data) + 10
        author_column_width = max(stringWidth(str(row[2]), font_name, font_size) for row in data) + 10
        date_create_column_width = max(stringWidth(str(row[3]), font_name, font_size) for row in data) + 10
        price_column_width = max(stringWidth(str(row[4]), font_name, font_size) for row in data) + 15
        procent_column_width = max(stringWidth(str(row[3]), font_name, font_size) for row in data) + 10
        price_procent_column_width = max(stringWidth(str(row[4]), font_name, font_size) for row in data) + 15


        return [
            number_column_width,
            name_column_width,
            author_column_width,
            date_create_column_width,
            price_column_width,
            procent_column_width,
            price_procent_column_width,
        ]

    def get_table_data(self, procurements: List['Procurement']) -> List[List[str]]:
        """Собирает и возвращает данные для таблицы."""
        headers = ["№", "Наименование", "Пользователь", "Дата", "Сумма", "%", "Сумма%"]
        data = [headers]

        for idx, procurement in enumerate(procurements, start=1):
            # Форматируем числа с разделителями тысяч и двумя знаками после запятой
            price = f"{getattr(procurement, 'price', 0):,.2f}".replace(",", "`").replace(".", ",") + " руб."
            price_procent = f"{getattr(procurement, 'price_procent', 0):,.2f}".replace(",", "`").replace(".", ",")

            data.append([
                str(idx),
                str(getattr(procurement, 'name', 'Имя не указанно')),
                str(getattr(procurement, 'author', 'Не указано')),
                f"{str(getattr(procurement, 'date_create', 0))}",
                price,
                f"{str(getattr(procurement, 'procent', 0))} %",
                price_procent,
            ])
        return data


class ProcurementViewXlsx(View):
    """View вывода XLSX закупок"""
    def get(self, request, *args, **kwargs) -> HttpResponse:
        slug = self.kwargs['slug']
        procurements = Procurement.objects.filter(complectation__slug=slug).select_related('complectation')

        # Проверяем, что есть данные
        if not procurements.exists():
            raise Http404("Записи не найдены для данного комплектации.")

        buffer = io.BytesIO()

        # Создаем рабочую книгу и активный лист
        wb = Workbook()
        ws = wb.active
        ws.title = "Procurements"

        # Устанавливаем заголовок
        title = f'Информация о услугах "{procurements.first().complectation}"'
        ws.merge_cells('A1:F1')  # Объединяем ячейки для заголовка
        ws['A1'] = title
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].font = Font(bold=True, size=14)

        # Определяем заголовки столбцов и добавляем данные без пустой строки
        headers = ["Название", "Пользователь", "Дата поступления", "Сумма", "Процент", "Сумма%"]
        ws.append(headers)
        for cell in ws[2]:  # Устанавливаем стиль только для строки с заголовками
            cell.font = Font(bold=True)

        # Заполняем данными из базы данных
        for procurement in procurements:
            ws.append([procurement.name, str(procurement.author), procurement.date_create, procurement.price, procurement.procent, procurement.price_procent])

        # Настраиваем ширину столбцов
        for col_num, column_title in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 50  # Ширина каждого столбца

        # Сохраняем рабочую книгу в буфер
        wb.save(buffer)
        buffer.seek(0)

        # Формируем уникальное имя файла с UUID
        unique_id = uuid.uuid4()
        filename = f'Procurements_Info_{unique_id}.xlsx'
        quoted_filename = quote(filename)

        # Формируем и отправляем ответ
        response = HttpResponse(buffer,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response
