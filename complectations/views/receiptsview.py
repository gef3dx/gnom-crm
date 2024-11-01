import os
import uuid
from typing import List
from django.conf import settings
from django.contrib import messages
from django.db.models import Sum
import datetime
from django.http import HttpResponse, Http404
from openpyxl.styles import Alignment, Font
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.platypus import Paragraph, Spacer
from reportlab.platypus import Image as Img
from django.views import View
from complectations.utils import (
    setup_pdf,
    setup_styles,
    create_table,
    get_filename,
)
import io
from django.urls import reverse
from django.shortcuts import get_object_or_404
from django.views.generic import (
    ListView,
    UpdateView,
    CreateView,
    DeleteView
)
from complectations.models import (
    Complectation,
    Receipts,
)
from complectations.forms import (
    ReceiptsForm,
)
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from urllib.parse import quote


class ReceiptsViewList(ListView):
    """View вывод поступлений средств в комплектацию"""
    template_name = 'complectations/receipt/receipts.html'
    model = Receipts
    context_object_name = 'receipts'

    def get_queryset(self):
        slug = self.kwargs['slug']
        queryset = Receipts.objects.filter(complectation__slug=slug).select_related('complectation')
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        slug = self.kwargs['slug']
        comp = get_object_or_404(Complectation, slug=slug)
        context['product_id'] = comp.id
        context['comp_name'] = comp.name
        context['comp_slug'] = comp.slug

        # Вычисление суммы и добавление в контекст
        total_sum = self.get_queryset().aggregate(total_sum=Sum('price'))
        context['sum'] = total_sum['total_sum']

        return context


class ReceiptsViewAdd(CreateView):
    """View добавления поступлений средств в комплектацию"""
    template_name = "complectations/receipt/receiptsadd.html"
    model = Receipts
    form_class = ReceiptsForm
    url = ""

    def form_valid(self, form, **kwargs):
        obj = form.save(commit=False)
        obj.author = self.request.user
        comp = Complectation.objects.filter(id=self.kwargs['product_id'])
        self.url = comp[0].slug
        obj.complectation = comp[0]
        messages.success(self.request, "Поступление добавлено!")
        return super(ReceiptsViewAdd, self).form_valid(form)

    def get_success_url(self):
        return reverse('receiptspage', kwargs={'slug': self.url})

    # def get(self, request, *args, **kwargs):
    #     if self.request.user.is_staff:
    #         self.object = None
    #         return super().get(request, *args, **kwargs)
    #     else:
    #         return redirect('home')
    #
    # def post(self, request, *args, **kwargs):
    #     if self.request.user.is_staff:
    #         self.object = None
    #         return super().post(request, *args, **kwargs)
    #     else:
    #         return redirect('home')


class ReceiptsViewEdit(UpdateView):
    """View редактирования поступлений средств в комплектацию"""
    template_name = "complectations/receipt/receiptsedit.html"
    model = Receipts
    form_class = ReceiptsForm

    def form_valid(self, form):
        messages.success(self.request, "Поступление сохранено!")
        return super().form_valid(form)

    def get_success_url(self):
        # Используем self.object, чтобы избежать повторного запроса к базе данных
        return reverse('receiptspage', kwargs={'slug': self.object.complectation.slug})

    # def get(self, request, *args, **kwargs):
    #     if self.request.user.is_staff:
    #         self.object = None
    #         return super().get(request, *args, **kwargs)
    #     else:
    #         return redirect('home')
    #
    # def post(self, request, *args, **kwargs):
    #     if self.request.user.is_staff:
    #         self.object = None
    #         return super().post(request, *args, **kwargs)
    #     else:
    #         return redirect('home')


class ReceiptsViewDelete(DeleteView):
    """View удаления поступлений средств в комплектацию"""
    model = Receipts
    template_name = "complectations/receipt/receiptsdelete.html"

    def form_valid(self, form):
        messages.success(self.request, "Поступление удаленно!")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('receiptspage', kwargs={'slug': self.object.complectation.slug})

    # def get(self, request, *args, **kwargs):
    #     if self.request.user.is_staff:
    #         self.object = None
    #         return super().get(request, *args, **kwargs)
    #     else:
    #         return redirect('home')
    #
    # def post(self, request, *args, **kwargs):
    #     if self.request.user.is_staff:
    #         self.object = None
    #         return super().post(request, *args, **kwargs)
    #     else:
    #         return redirect('home')


class ReceiptsViewPDF(View):
    """View вывода PDF поступлений"""
    def get(self, request, *args, **kwargs) -> HttpResponse:
        slug = self.kwargs['slug']
        receipts = Receipts.objects.filter(complectation__slug=slug).select_related('complectation')

        # Проверяем, что есть данные
        if not receipts.exists():
            raise Http404("Записи не найдены для данного комплектации.")

        buffer = io.BytesIO()

        # Настройка PDF и стили
        pdf = setup_pdf(buffer)
        styles = setup_styles()
        data = self.get_table_data(receipts)
        col_widths = self.calculate_column_widths(data)
        table = create_table(data, col_widths)

        # Логотип
        logo_path = os.path.join(settings.BASE_DIR, "static", "img", "gnomlogob.png")
        logo = Img(logo_path, width=2 * inch, height=1 * inch)

        # Стиль заголовка по центру
        styles.add(ParagraphStyle(name="CenteredHeading1", parent=styles["Heading1"], alignment=TA_CENTER))
        styles.add(ParagraphStyle(name="CenteredHeading2", parent=styles["Heading1"], alignment=TA_CENTER, fontSize=14))

        # Вычисление суммы и добавление в контекст
        total_sum = receipts.aggregate(total_sum=Sum('price'))["total_sum"]

        # Элементы документа
        elements = [
            logo, Spacer(1, 0.2 * inch),
            Paragraph(f'Информация о поступлениях "{receipts.first().complectation}"', styles["CenteredHeading1"]),
            Paragraph(f'Сумма всех поступлений: {total_sum:,.2f} р на {str(datetime.datetime.now())[:11]}'.replace(",", "`").replace(".", ","), styles["CenteredHeading2"]),
            Spacer(1, 0.2 * inch), table
        ]
        pdf.build(elements)

        # Возвращаем буфер к началу
        buffer.seek(0)

        # Формируем ответ
        filename = get_filename(receipts)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    def calculate_column_widths(self, data: List[List[str]]) -> List[float]:
        """Вычисляет и возвращает ширины колонок для таблицы."""
        font_name = 'FreeSans'
        font_size = 10
        number_column_width = max(stringWidth(str(row[0]), font_name, font_size) for row in data) + 10
        name_column_width = max(stringWidth(str(row[1]), font_name, font_size) for row in data) + 10
        author_column_width = max(stringWidth(str(row[2]), font_name, font_size) for row in data) + 10
        date_create_column_width = max(stringWidth(str(row[3]), font_name, font_size) for row in data) + 10
        price_column_width = max(stringWidth(str(row[4]), font_name, font_size) for row in data) + 15

        return [
            number_column_width,
            name_column_width,
            author_column_width,
            date_create_column_width,
            price_column_width,
        ]

    def get_table_data(self, receipts: List['Receipts']) -> List[List[str]]:
        """Собирает и возвращает данные для таблицы."""
        headers = ["№", "Наименование", "Пользователь", "Дата поступления", "Сумма поступления"]
        data = [headers]

        for idx, receipt in enumerate(receipts, start=1):
            price = f"{getattr(receipt, 'price', 0):,.2f}".replace(",", "`").replace(".", ",") + " руб."
            data.append([
                str(idx),
                str(getattr(receipt, 'name', 'Имя не указанно')),
                str(getattr(receipt, 'author', 'Не указано')),
                f"{str(getattr(receipt, 'date_create', 0))}",
                price,
            ])
        return data


class ReceiptsViewXlsx(View):
    """View вывода XLSX поступлений"""
    def get(self, request, *args, **kwargs) -> HttpResponse:
        slug = self.kwargs['slug']
        receipts = Receipts.objects.filter(complectation__slug=slug).select_related('complectation')

        # Проверяем, что есть данные
        if not receipts.exists():
            raise Http404("Записи не найдены для данного комплектации.")

        buffer = io.BytesIO()

        # Создаем рабочую книгу и активный лист
        wb = Workbook()
        ws = wb.active
        ws.title = "Receipts"

        # Устанавливаем заголовок
        title = f'Информация о поступлениях "{receipts.first().complectation}"'
        ws.merge_cells('A1:D1')  # Объединяем ячейки для заголовка
        ws['A1'] = title
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].font = Font(bold=True, size=14)

        # Определяем заголовки столбцов и добавляем данные без пустой строки
        headers = ["Название", "Пользователь", "Дата поступления", "Цена за единицу"]
        ws.append(headers)
        for cell in ws[2]:  # Устанавливаем стиль только для строки с заголовками
            cell.font = Font(bold=True)

        # Заполняем данными из базы данных
        for receipt in receipts:
            ws.append([receipt.name, str(receipt.author), receipt.date_create, receipt.price])

        # Настраиваем ширину столбцов
        for col_num, column_title in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 50  # Ширина каждого столбца

        # Сохраняем рабочую книгу в буфер
        wb.save(buffer)
        buffer.seek(0)

        # Формируем уникальное имя файла с UUID
        unique_id = uuid.uuid4()
        filename = f'Receipts_Info_{unique_id}.xlsx'
        quoted_filename = quote(filename)

        # Формируем и отправляем ответ
        response = HttpResponse(buffer,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response
