import os
import uuid
from typing import List
from django.conf import settings
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.db.models import Sum
import datetime
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.platypus import Paragraph, Spacer
from reportlab.platypus import Image as Img
from django.views import View
from complectations.utils import (
    compress_image,
    setup_pdf,
    setup_styles,
    create_table,
    get_filename,
)
from PIL import Image
import io
from django.urls import reverse
from django.utils.timezone import now
from django.shortcuts import get_object_or_404
from django.views.generic import (
    ListView,
    UpdateView,
    CreateView,
    DeleteView
)
from django.views.generic.detail import DetailView
from complectations.models import (
    Complectation,
    GroupProduct,
    Product,
)
from complectations.forms import (
    ProductForm,
)
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from urllib.parse import quote


class ProductViewList(ListView):
    """View для вывода продуктов комплектации"""
    model = Product
    template_name = 'complectations/product/product.html'
    context_object_name = 'products'

    def get_queryset(self):
        slug = self.kwargs['slug']
        # Получаем набор продуктов для данной комплектации
        return Product.objects.filter(complectation__slug=slug)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Получаем объект комплектации только один раз
        complectation = get_object_or_404(Complectation, slug=self.kwargs['slug'])

        # Подготавливаем заранее набор продуктов
        products = self.get_queryset()

        # Агрегируем нужные суммы за один запрос
        totals = products.aggregate(
            total_sum=Sum('sum_price_count'),
            total_prepayment=Sum('prepayment'),
            total_remains=Sum('remains')
        )

        # Контекстные данные
        context.update({
            'product_id': complectation.id,
            'product_name': complectation.name,
            'product_slug': complectation.slug,
            'sum': totals['total_sum'],  # Общая сумма
            'prepayment': totals['total_prepayment'],  # Общая предоплата
            'remains': totals['total_remains'],  # Остаток
            'date_now': now().date(),  # Текущая дата
            'date_plus': now().date() + datetime.timedelta(days=7),  # Дата через 7 дней
            'groups': GroupProduct.objects.all(),  # Все группы продуктов
        })

        return context


class ProductViewListGroup(ListView):
    """View для вывода продуктов комплектации по группе"""
    template_name = 'complectations/product/product.html'
    model = Product
    context_object_name = 'products'

    def get_queryset(self):
        # Фильтруем продукты по комплектации и группе
        slug = self.kwargs['slug']
        group_slug = self.kwargs['group']
        return Product.objects.filter(complectation__slug=slug, group__slug=group_slug)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Получаем объект Complectation только один раз
        complectation = get_object_or_404(Complectation, slug=self.kwargs['slug'])
        products = self.get_queryset()

        # Агрегируем нужные суммы в одном запросе
        totals = products.aggregate(
            total_sum=Sum('sum_price_count'),
            total_prepayment=Sum('prepayment'),
            total_remains=Sum('remains')
        )

        # Заполняем контекстные данные
        context.update({
            'product_id': complectation.id,
            'product_name': complectation.name,
            'product_slug': complectation.slug,
            'sum': totals['total_sum'],  # Общая сумма
            'prepayment': totals['total_prepayment'],  # Общая предоплата
            'remains': totals['total_remains'],  # Остаток
            'date_now': now().date(),  # Текущая дата
            'date_plus': now().date() + datetime.timedelta(days=7),  # Дата через 7 дней
            'groups': GroupProduct.objects.all(),  # Все группы продуктов
        })

        # Устанавливаем grouppdf и grouppdfslug из первого продукта, если он есть
        if products.exists():
            first_product = products.first()
            context['grouppdf'] = first_product.group
            context['grouppdfslug'] = first_product.group.slug
        else:
            context['grouppdf'] = ""
            context['grouppdfslug'] = ""

        return context


class ProductViewDetale(DetailView):
    """View для просмотра одной записи продукта"""
    model = Product
    template_name = 'complectations/product/productdetale.html'
    context_object_name = 'product'


class ProductViewAdd(CreateView):
    """View для добавления продукта"""
    template_name = "complectations/product/productadd.html"
    model = Product
    form_class = ProductForm

    def form_valid(self, form):
        obj = form.save(commit=False)
        obj.author = self.request.user
        comp = get_object_or_404(Complectation, id=self.kwargs['product_id'])
        obj.complectation = comp
        obj.sum_price_count = obj.count * obj.price
        obj.remains = obj.prepayment - obj.sum_price_count

        image_field = self.request.FILES.get('image')

        if image_field:
            allowed_extensions = ['jpg', 'jpeg', 'png']
            file_extension = image_field.name.split('.')[-1].lower()
            if file_extension not in allowed_extensions:
                form.add_error('image', "Только файлы с расширениями: jpg, jpeg, png допустимы.")
                return self.form_invalid(form)

            if image_field.size > 2 * 1024 * 1024:
                form.add_error('image', "Размер изображения не должен превышать 2 МБ.")
                return self.form_invalid(form)

            # Открываем и сжимаем изображение
            img = Image.open(image_field)
            compressed_image = compress_image(img)

            # Генерируем уникальное имя для файла
            unique_filename = f"img_{uuid.uuid4().hex}.jpg"

            # Определяем путь для сохранения изображения
            slug_folder = os.path.join('product', comp.slug)  # относительный путь
            full_path = os.path.join(settings.MEDIA_ROOT, slug_folder)  # абсолютный путь
            os.makedirs(full_path, exist_ok=True)  # Создаем папку, если ее нет

            # Сохраняем изображение в нужной папке
            output = io.BytesIO()
            compressed_image.save(output, format='JPEG')
            output.seek(0)

            # Задаем относительный путь для поля image
            obj.image = InMemoryUploadedFile(
                output, 'ImageField', os.path.join(slug_folder, unique_filename),
                'image/jpeg', output.tell(), None
            )

        obj.save()
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('productpage', kwargs={'slug': self.object.complectation.slug})


class ProductViewEdit(UpdateView):
    """View для редактирования продукта"""
    template_name = "complectations/product/productedit.html"
    model = Product
    form_class = ProductForm

    def form_valid(self, form):
        obj = form.save(commit=False)
        image_field = self.request.FILES.get('image')

        if image_field:
            # Проверка расширения
            allowed_extensions = ['jpg', 'jpeg', 'png']
            file_extension = image_field.name.split('.')[-1].lower()
            if file_extension not in allowed_extensions:
                form.add_error('image', "Только файлы с расширениями: jpg, jpeg, png допустимы.")
                return self.form_invalid(form)

            # Проверка размера файла
            if image_field.size > 2 * 1024 * 1024:
                form.add_error('image', "Размер изображения не должен превышать 2 МБ.")
                return self.form_invalid(form)

            # Открываем и сжимаем изображение
            img = Image.open(image_field)
            compressed_image = compress_image(img)

            # Сохраняем сжатое изображение
            output = io.BytesIO()
            compressed_image.save(output, format='JPEG')
            output.seek(0)

            # Обновляем поле image для модели с новым изображением
            obj.image = InMemoryUploadedFile(
                output, 'ImageField', image_field.name,
                'image/jpeg', output.tell(), None
            )

        obj.save()
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('productpage', kwargs={'slug': self.object.complectation.slug})


class ProductViewDelete(DeleteView):
    """View для удаления продукта"""
    template_name = "complectations/product/productdelete.html"
    model = Product

    def get_success_url(self):
        return reverse('productpage', kwargs={'slug': self.object.complectation.slug})


class ProductViewPDF(View):
    """View вывода PDF комплектации"""
    def get(self, request, *args, **kwargs) -> HttpResponse:
        products = Product.objects.all()
        buffer = io.BytesIO()

        # Используем функции из utils.py
        pdf = setup_pdf(buffer)
        styles = setup_styles()
        data = self.get_table_data(products)
        col_widths = self.calculate_column_widths(data)
        table = create_table(data, col_widths)

        # Настраиваем логотип
        logo_path = os.path.join(settings.BASE_DIR, "static", "img", "gnomlogob.png")
        logo = Img(logo_path, width=2 * inch, height=1 * inch)

        styles = setup_styles()
        styles.add(ParagraphStyle(name="CenteredHeading1", parent=styles["Heading1"], alignment=TA_CENTER))

        # Создание элементов PDF-документа
        elements = [
            logo, Spacer(1, 0.2 * inch),
            Paragraph(f'Информация о продуктах "{products.first().complectation}"', styles["CenteredHeading1"]),
            Spacer(1, 0.2 * inch), table
        ]
        pdf.build(elements)

        # Переместим указатель буфера в начало, чтобы данные можно было считать
        buffer.seek(0)

        # Формируем и отправляем ответ
        filename = get_filename(products)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    def calculate_column_widths(self, data: List[List[str]]) -> List[float]:
        """Вычисляет и возвращает ширины колонок для таблицы."""
        font_name = 'FreeSans'
        font_size = 10
        number_column_width = max(stringWidth(str(row[0]), font_name, font_size) for row in data) + 10
        name_column_width = max(stringWidth(str(row[1]), font_name, font_size) for row in data) + 10
        group_column_width = max(stringWidth(str(row[2]), font_name, font_size) for row in data) + 10
        quantity_column_width = max(stringWidth(str(row[3]), font_name, font_size) for row in data) + 10
        price_column_width = max(stringWidth(str(row[4]), font_name, font_size) for row in data) + 15
        total_price_column_width = max(stringWidth(str(row[5]), font_name, font_size) for row in data) + 15

        return [
            number_column_width,
            name_column_width,
            group_column_width,
            quantity_column_width,
            price_column_width,
            total_price_column_width
        ]

    def get_table_data(self, products: List['Product']) -> List[List[str]]:
        """Собирает и возвращает данные для таблицы."""
        headers = ["№", "Имя продукта", "Группа", "Количество", "Цена за единицу", "Общая стоимость"]
        data = [headers]

        for idx, product in enumerate(products, start=1):
            data.append([
                str(idx),
                str(getattr(product, 'name', 'Имя не указанно')),
                str(getattr(product, 'group', 'Не указано')),
                f"{str(getattr(product, 'count', 0))} {str(getattr(product, 'meas', ''))}",
                f"{str(getattr(product, 'price', 0))} руб.",
                f"{str(getattr(product, 'sum_price_count', 0))} руб.",
            ])
        return data


class ProductViewXlsx(View):
    """View вывода XLSX комплектации"""
    def get(self, request, *args, **kwargs) -> HttpResponse:
        products = Product.objects.all()
        buffer = io.BytesIO()

        # Создаем рабочую книгу и активный лист
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        # Устанавливаем заголовок
        title = f'Информация о продуктах "{products.first().complectation}"'
        ws.merge_cells('A1:E1')  # Объединяем ячейки для заголовка
        ws['A1'] = title
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].font = Font(bold=True, size=14)

        # Определяем заголовки столбцов и добавляем данные без пустой строки
        headers = ["Название", "Группа", "Количество", "Цена за единицу", "Общая стоимость"]
        ws.append(headers)
        for cell in ws[2]:  # Устанавливаем стиль только для строки с заголовками
            cell.font = Font(bold=True)

        # Заполняем данными из базы данных
        for product in products:
            ws.append([product.name, str(product.group), product.count, product.price, product.sum_price_count])

        # Настраиваем ширину столбцов
        for col_num, column_title in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20  # Ширина каждого столбца

        # Сохраняем рабочую книгу в буфер
        wb.save(buffer)
        buffer.seek(0)

        # Формируем уникальное имя файла с UUID
        unique_id = uuid.uuid4()
        filename = f'Product_Info_{products.first().complectation}_{unique_id}.xlsx'
        quoted_filename = quote(filename)

        # Формируем и отправляем ответ
        response = HttpResponse(buffer,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response